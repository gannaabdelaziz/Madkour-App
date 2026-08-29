"""Extract a populated, real region of the generated workbook (values + fills)
so the preview page shows genuine output rather than mocked-up numbers."""
import json
import datetime
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension

_orig = ColumnDimension.__init__
ColumnDimension.__init__ = lambda s, *a, **k: (k.pop("level", None), _orig(s, *a, **k))[1]

SRC = r"D:/Ganna/The Cashflow Automation/Madkour-App/demo/output/DEMO_Cashflow_Baseline.xlsx"
OUT = r"D:/Ganna/The Cashflow Automation/Madkour-App/demo/grid.json"
SHEET = "DEMO 220-66-11 kV Cash out USD"

COLS = [1, 2, 5, 6, 25, 26]
HEADERS = ["Parent Task", "Start ID", "Item Description", "Dry Cost USD",
           "Prog. Start", "Prog. End"]
WIDTHS = [96, 122, 232, 116, 84, 84]
ROWS = list(range(4, 30))

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[SHEET]


def argb(c):
    try:
        if c and isinstance(c.rgb, str) and len(c.rgb) == 8:
            return "#" + c.rgb[2:]
    except Exception:
        pass
    return None


grid = []
for r in ROWS:
    row = []
    for c in COLS:
        cell = ws.cell(r, c)
        v = cell.value
        if isinstance(v, datetime.datetime):
            txt, num = v.strftime("%b-%y"), False
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            txt, num = f"{v:,.0f}", True
        else:
            txt, num = ("" if v is None else str(v)), False
        bg = None
        try:
            if cell.fill and cell.fill.patternType:
                bg = argb(cell.fill.fgColor)
        except Exception:
            pass
        row.append({"v": txt, "num": num, "bg": bg,
                    "fg": argb(cell.font.color) if cell.font else None,
                    "b": bool(cell.font and cell.font.bold)})
    grid.append(row)

json.dump({"sheets": wb.sheetnames, "active": SHEET, "headers": HEADERS,
           "widths": WIDTHS, "grid": grid},
          open(OUT, "w", encoding="utf8"))
print("wrote", OUT)
print("sheets:", len(wb.sheetnames), "| rows:", len(grid))
print("sample:", [c["v"] for c in grid[1]])
