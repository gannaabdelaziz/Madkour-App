"""Anonymize the Primavera baseline: shift every date by a fixed offset so the
schedule no longer maps to the real project. Activity IDs and generic
construction activity names are kept so the demo parses identically.
"""
import openpyxl
import datetime

SRC = r"D:/Ganna/The Cashflow Automation/PS3/inputs/Primavera Baseline.xlsx"
OUT = r"D:/Ganna/The Cashflow Automation/Madkour-App/demo/inputs/Primavera Baseline.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["Activities"]

# find earliest date across both date columns
earliest = None
for r in range(2, ws.max_row + 1):
    for c in (3, 4):
        v = ws.cell(r, c).value
        if isinstance(v, datetime.datetime):
            if earliest is None or v < earliest:
                earliest = v

target = datetime.datetime(2026, 1, 5)
offset = target - earliest
print("earliest:", earliest.date(), "-> shifting by", offset.days, "days")

n = 0
for r in range(2, ws.max_row + 1):
    for c in (3, 4):
        cell = ws.cell(r, c)
        if isinstance(cell.value, datetime.datetime):
            cell.value = cell.value + offset
            n += 1

wb.save(OUT)
print("shifted", n, "dates ->", OUT)
