"""Build anonymized demo copies of the three cashflow input files.

Money is scaled by ONE factor across all three files so every cross-file
total still reconciles. Percentages, durations and exchange rates are left
alone. Hardcoded total rows are recomputed from the scaled line items.
"""
import openpyxl
import datetime
import os

SRC = r"D:/Ganna/The Cashflow Automation/Madkour-App/uploads"
OUT = r"D:/Ganna/The Cashflow Automation/Madkour-App/demo/inputs"
SCALE = 0.7362  # single uniform factor

os.makedirs(OUT, exist_ok=True)


def scale(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(v * SCALE, 2)
    return v


# ---------------------------------------------------------------- BOQ
wb = openpyxl.load_workbook(os.path.join(SRC, "BOQ for Cash IN.xlsx"))
ws = wb["BOQ"]
PRICE_COLS = (7, 8, 9)  # USD, EUR, EGP
totals = {c: 0.0 for c in PRICE_COLS}
for r in range(6, ws.max_row + 1):
    for c in PRICE_COLS:
        cell = ws.cell(r, c)
        if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
            cell.value = scale(cell.value)
            totals[c] += cell.value
# row 5 holds the hardcoded column totals
for c in PRICE_COLS:
    ws.cell(5, c).value = round(totals[c], 2)
wb.save(os.path.join(OUT, "BOQ for Cash IN.xlsx"))
print("BOQ totals ->", {k: round(v) for k, v in totals.items()})

# ------------------------------------------------------------- Budget
wb = openpyxl.load_workbook(os.path.join(SRC, "Budget For Cash Out.xlsx"))
ws = wb["Break Down"]
AMOUNT_COL = 11
n = 0
for r in range(4, ws.max_row + 1):
    cell = ws.cell(r, AMOUNT_COL)
    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        cell.value = scale(cell.value)
        n += 1
wb.save(os.path.join(OUT, "Budget For Cash Out.xlsx"))
print("Budget amounts scaled:", n)

# -------------------------------------------------------- Assumptions
wb = openpyxl.load_workbook(os.path.join(SRC, "PS3_Assumptions.xlsx"))
ws = wb["Assumptions"]
ws.cell(4, 3).value = "DEMO 220/66/11 kV GIS Substation"
ws.cell(5, 3).value = datetime.datetime(2026, 1, 1)
# row 13 = contract value per currency, must match scaled BOQ totals
for col, src in zip((2, 3, 4), PRICE_COLS):
    ws.cell(13, col).value = round(totals[src], 2)
wb.save(os.path.join(OUT, "Assumptions.xlsx"))
print("Assumptions: project renamed, contract value re-synced")
print("Wrote ->", OUT)
